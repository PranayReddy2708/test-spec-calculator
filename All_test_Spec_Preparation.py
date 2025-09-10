import streamlit as st
import pandas as pd
import json
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import sqlite3

# --- Material Data ---
Material_detailes = {
    "Steel": {"Fatigue constant": 3, "youngs Modulus": 0.205},
    "Aluminum": {"Fatigue constant": 5, "youngs Modulus": 0.07},
}

# --- Calculator Functions (Unchanged) ---
def panic_brake_calculation(inputs):
    force = inputs["Mass of the Vehicle (kg)"] * inputs["Max Deceleration (m/s^2)"]
    torque = force * inputs["Tyre rolling radius (m)"]
    applied_force = (torque / inputs["Fixture arm length (m)"]) / 9.81
    total_cycles = inputs["Total life (km)"] / inputs["Road to rig factor"]
    return {"Required Load (kg)": applied_force, "Required Cycles": total_cycles}

def Front_Fork_Fatigue_calculation(inputs):
    Max_BM = inputs["fork Length (mm)"] * inputs["Max Load (kgf)"]
    Min_BM = inputs["fork Length (mm)"] * inputs["Min Load (kgf)"]
    Max_strain = (Max_BM * inputs["Calibration factor"]) + inputs["Calibration constant"]
    Min_strain = (Min_BM * inputs["Calibration factor"]) + inputs["Calibration constant"]
    Material_type = inputs["Material"]
    Material_data = Material_detailes[Material_type]
    Max_stress = Max_strain * Material_data["youngs Modulus"]
    Min_stress = Min_strain * Material_data["youngs Modulus"]
    Mean_stress = (Max_stress + Min_stress) / 2
    Amplitude_stress = (Max_stress - Min_stress) / 2
    Mean_corrected_stress = Amplitude_stress / (1 - (Mean_stress / Amplitude_stress))
    Damage_per_cycle = Mean_corrected_stress ** Material_data["Fatigue constant"]
    Number_of_Cycles = (inputs["Target Damage"] * inputs["Factor of Safety"]) / Damage_per_cycle
    return {"Total Number of cycles": Number_of_Cycles}

# --- Test Definitions (Unchanged) ---
test_definitions = {
    "Panic Brake Fatigue": {
        "inputs": {
            "Max Deceleration (m/s^2)": {"type": "number", "min_value": 1.0, "value": 9.8},
            "Mass of the Vehicle (kg)": {"type": "number", "min_value": 1.0, "value": 250.0},
            "Tyre rolling radius (m)": {"type": "number", "value": 0.3},
            "Fixture arm length (m)": {"type": "number", "value": 0.5},
            "Total life (km)": {"type": "number", "min_value": 1.0, "value": 100000.0},
            "Road to rig factor": {"type": "number", "value": 100.0},
        },
        "calculator": panic_brake_calculation
    },
    "Front Fork Fatigue": {
        "inputs": {
            "Target Damage": {"type": "number", "value": 1.0},
            "fork Length (mm)": {"type": "number", "value": 400.0},
            "Max Load (kgf)": {"type": "number", "value": 50.0},
            "Min Load (kgf)": {"type": "number", "value": -20.0},
            "Calibration factor": {"type": "text", "value": "0.00054", "placeholder": "Enter: 0.00054"},
            "Calibration constant": {"type": "text", "value": "-1.356", "placeholder": "Enter: -1.356"},
            "Material": {"type": "selectbox", "options": list(Material_detailes.keys())},
            "Factor of Safety": {"type": "number", "value": 2.0},
        },
        "calculator": Front_Fork_Fatigue_calculation
    }
}

# --- Database Functions (Updated) ---
def init_db():
    conn = sqlite3.connect('test_spec_history.db')
    c = conn.cursor()
    # Added project_code column
    c.execute('''
        CREATE TABLE IF NOT EXISTS calculations (
            id INTEGER PRIMARY KEY,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            project_code TEXT,
            test_name TEXT NOT NULL,
            inputs TEXT NOT NULL,
            results TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

def save_calculation(project_code, test_name, inputs, results):
    conn = sqlite3.connect('test_spec_history.db')
    c = conn.cursor()
    # Insert project_code into the database
    c.execute(
        "INSERT INTO calculations (project_code, test_name, inputs, results) VALUES (?, ?, ?, ?)",
        (project_code, test_name, json.dumps(inputs), json.dumps(results))
    )
    conn.commit()
    conn.close()

def get_history():
    conn = sqlite3.connect('test_spec_history.db')
    # Select project_code to display in history
    query = "SELECT id, timestamp, project_code, test_name FROM calculations ORDER BY timestamp DESC"
    history_df = pd.read_sql_query(query, conn)
    conn.close()
    return history_df

def load_calculation_from_db(calc_id):
    conn = sqlite3.connect('test_spec_history.db')
    c = conn.cursor()
    c.execute("SELECT inputs, results FROM calculations WHERE id = ?", (calc_id,))
    row = c.fetchone()
    conn.close()
    if row:
        return json.loads(row[0]), json.loads(row[1])
    return None, None

init_db()

# --- Report Generation Functions (Unchanged) ---
def create_excel_report(test_name, inputs, results):
    """Create a professional Excel report with formatting"""
    output = BytesIO()
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Test Specification Report"

    # Define styles
    header_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    regular_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Header Section
    ws.merge_cells("A1:D1")
    ws["A1"] = "TEST SPECIFICATION CALCULATION REPORT"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # Test Information Section
    row = 3
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "TEST INFORMATION"
    ws[f"A{row}"].font = subheader_font
    ws[f"A{row}"].fill = subheader_fill
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    row += 1
    test_info = [
        ["Test Type:", test_name],
        ["Calculation Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Generated By:", "Automated Test Spec Calculator"],
        ["Version:", "2.0"]
    ]
    for info in test_info:
        ws[f"A{row}"] = info[0]
        ws[f"A{row}"].font = bold_font
        ws[f"B{row}"] = info[1]
        ws[f"B{row}"].font = regular_font
        row += 1

    # Input Parameters Section
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "INPUT PARAMETERS"
    ws[f"A{row}"].font = subheader_font
    ws[f"A{row}"].fill = subheader_fill
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    row += 1
    ws[f"A{row}"] = "Parameter"
    ws[f"B{row}"] = "Value"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].font = bold_font
    row += 1
    for param, value in inputs.items():
        ws[f"A{row}"] = param
        ws[f"B{row}"] = str(value)
        row += 1

    # Results Section
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "CALCULATION RESULTS"
    ws[f"A{row}"].font = subheader_font
    ws[f"A{row}"].fill = subheader_fill
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    row += 1
    ws[f"A{row}"] = "Result Parameter"
    ws[f"B{row}"] = "Calculated Value"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{row}"].font = bold_font
    row += 1
    for param, value in results.items():
        ws[f"A{row}"] = param
        ws[f"B{row}"] = f"{value:.6f}" if isinstance(value, float) else str(value)
        ws[f"D{row}"] = "✓ Calculated"
        row += 1

    # Apply borders and adjust widths
    for r in ws.iter_rows(min_row=1, max_row=row-1, min_col=1, max_col=4):
        for cell in r:
            cell.border = border
    for col_letter in ["A", "B", "C", "D"]:
        ws.column_dimensions[col_letter].width = 25

    workbook.save(output)
    return output.getvalue()

def create_pdf_report(test_name, inputs, results):
    """Create a professional PDF report"""

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
    story = []

    # Get styles
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Center alignment
        textColor=colors.darkblue
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.darkblue,
        borderWidth=1,
        borderColor=colors.darkblue,
        borderPadding=5,
        backColor=colors.lightgrey
    )

    # Title
    title = Paragraph("TEST SPECIFICATION CALCULATION REPORT", title_style)
    story.append(title)
    story.append(Spacer(1, 20))

    # Test Information Section
    story.append(Paragraph("TEST INFORMATION", heading_style))

    test_info_data = [
        ["Test Type:", test_name],
        ["Calculation Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Generated By:", "Automated Test Spec Calculator"],
        ["Version:", "2.0"]
    ]

    test_info_table = Table(test_info_data, colWidths=[2*inch, 3*inch])
    test_info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
    ]))

    story.append(test_info_table)
    story.append(Spacer(1, 20))

    # Input Parameters Section
    story.append(Paragraph("INPUT PARAMETERS", heading_style))

    input_data = [["Parameter", "Value", "Unit"]]
    for param, value in inputs.items():
        if isinstance(value, float):
            formatted_value = f"{value:.6f}"
        else:
            formatted_value = str(value)
        input_data.append([param, formatted_value, ""])  # Unit column can be enhanced

    input_table = Table(input_data, colWidths=[2.5*inch, 1.5*inch, 1*inch])
    input_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    ]))

    story.append(input_table)
    story.append(Spacer(1, 20))

    # Results Section
    story.append(Paragraph("CALCULATION RESULTS", heading_style))

    results_data = [["Result Parameter", "Calculated Value", "Unit", "Status"]]
    for param, value in results.items():
        if isinstance(value, float):
            formatted_value = f"{value:.6f}"
        else:
            formatted_value = str(value)
        results_data.append([param, formatted_value, "", "✓ Calculated"])

    results_table = Table(results_data, colWidths=[2*inch, 1.5*inch, 0.8*inch, 1.2*inch])
    results_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
    ]))

    story.append(results_table)
    story.append(Spacer(1, 30))

    # Footer
    footer_text = f"""
    <para alignment="center">
    <font size="8" color="grey">
    Generated by Automated Test Spec Calculator | 
    Report Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | 
    Page 1 of 1
    </font>
    </para>
    """
    story.append(Paragraph(footer_text, styles['Normal']))

    # Build PDF
    doc.build(story)
    return buffer.getvalue()
    
# === Streamlit UI ===
st.set_page_config(page_title="Test Spec Calculator", page_icon="🔧", layout="wide")
st.title("🔧 Automated Test Spec Calculation Tool")
st.markdown("---")

if 'inputs' not in st.session_state:
    st.session_state['inputs'] = {}

def update_input_state(label, key):
    st.session_state.inputs[label] = st.session_state[key]

selected_test = st.selectbox("🎯 Select Test Type", list(test_definitions.keys()))
test_config = test_definitions[selected_test]

col1, col2 = st.columns([2, 1])

with col1:
    # Project Code Input Field
    project_code = st.text_input("🏷️ Project Code / Name", placeholder="Enter a project identifier...")
    st.markdown("---")
    
    st.subheader(f"📋 Input Parameters for: {selected_test}")
    inputs = {}
    
    for label, config in test_config["inputs"].items():
        key = f"{selected_test}_{label}"
        current_value = st.session_state.inputs.get(label, config.get("value"))

        if config["type"] == "number":
            inputs[label] = st.number_input(label, key=key, value=current_value, on_change=update_input_state, args=(label, key))
        elif config["type"] == "selectbox":
            options = config["options"]
            index = options.index(current_value) if current_value in options else 0
            inputs[label] = st.selectbox(label, options, key=key, index=index, on_change=update_input_state, args=(label, key))
        elif config["type"] == "text":
            text_val = st.text_input(label, value=str(current_value), placeholder=config.get("placeholder",""), key=key, on_change=update_input_state, args=(label, key))
            try:
                inputs[label] = float(text_val) if text_val else 0.0
            except ValueError:
                st.error(f"❌ Enter a valid number for {label}")
                inputs[label] = 0.0

    calculate_clicked = st.button("🚀 Calculate", type="primary", use_container_width=True)

with col2:
    st.subheader("📊 Results & Export")

    if calculate_clicked:
        try:
            result = test_config["calculator"](inputs)
            st.session_state['last_calculation'] = {
                'project_code': project_code,
                'test_name': selected_test,
                'inputs': inputs,
                'results': result,
                'timestamp': datetime.now()
            }
        except Exception as e:
            st.error(f"❌ Calculation Error: {str(e)}")
            st.session_state.pop('last_calculation', None)

    if 'last_calculation' in st.session_state:
        calc_data = st.session_state['last_calculation']
        st.success("✅ Calculation Complete!")
        st.info(f"Project: **{calc_data['project_code']}**")
        for k, v in calc_data['results'].items():
            st.metric(label=k, value=f"{v:,.2f}" if isinstance(v, float) else str(v))

        if st.button("💾 Save Calculation to History", use_container_width=True):
            save_calculation(calc_data['project_code'], calc_data['test_name'], calc_data['inputs'], calc_data['results'])
            st.success("Calculation saved to history!")

st.markdown("---")

# --- Download Section (Unchanged) ---
# --- Download Section ---
if 'last_calculation' in st.session_state:
    st.subheader("📥 Export Report")
    calc_data = st.session_state['last_calculation']
    
    # Create columns for the download buttons
    d_col1, d_col2 = st.columns(2)

    with d_col1:
        excel_data = create_excel_report(calc_data['test_name'], calc_data['inputs'], calc_data['results'])
        st.download_button(
            label="📊 Download Excel Report",
            data=excel_data,
            file_name=f"{calc_data['project_code']}_{calc_data['test_name']}_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    with d_col2:
        pdf_data = create_pdf_report(calc_data['test_name'], calc_data['inputs'], calc_data['results'])
        st.download_button(
            label="📄 Download PDF Report",
            data=pdf_data,
            file_name=f"{calc_data['project_code']}_{calc_data['test_name']}_Report.pdf",
            mime="application/pdf",
            use_container_width=True
        )

# --- History Section ---
st.markdown("---")
with st.expander("📜 View Calculation History", expanded=False):
    history_df = get_history()

    if not history_df.empty:
        st.dataframe(history_df, use_container_width=True, hide_index=True)
        
        h_col1, h_col2 = st.columns([1, 2])
        with h_col1:
            calc_id_to_load = st.selectbox("Select Calculation ID to Load", options=history_df['id'])
        with h_col2:
            st.write("") 
            st.write("")
            if st.button("📥 Load Selected Calculation", key="load_hist"):
                inputs_from_db, _ = load_calculation_from_db(calc_id_to_load)
                if inputs_from_db:
                    st.session_state.inputs.update(inputs_from_db)
                    st.success(f"Loaded calculation #{calc_id_to_load}. Input fields updated.")
                    st.rerun()
                else:
                    st.error("Could not find the selected calculation.")
    else:
        st.info("No calculations have been saved yet.")

# --- Footer ---
st.markdown("---")
st.markdown("*Built with ❤️ using Streamlit | Version 4.0 with Project Codes*")
